# utils for RNAseq
import sys, json
from time import sleep
import random
random.seed(0)
from collections import OrderedDict

import numpy as np
import openpyxl as px
import xlwt
import requests

from sklearn.decomposition import PCA
from sklearn import manifold
from scipy.stats.mstats import zscore
import matplotlib.pyplot as plt
import scipy.spatial.distance as dist
import scipy.cluster.hierarchy as sch
from mpl_toolkits.mplot3d import Axes3D
from mpl_toolkits.mplot3d import proj3d
from matplotlib.lines import Line2D
from matplotlib import rcParams
rcParams['pdf.fonttype'] = 42 ## Output Type 3 (Type3) or Type 42 (TrueType)
rcParams['font.sans-serif'] = 'Arial'

from plots import (COLORS10, COLORS20, COLORS20b, enlarge_tick_fontsize)

ENRICHR_URL = 'http://amp.pharm.mssm.edu/Enrichr'

def _enrichr_add_list(genes, meta=''):
	"""POST a gene list to Enrichr server and return the list ids"""
	genes_str = '\n'.join(genes)
	payload = {
		'list': (None, genes_str),
		'description': (None, meta)
	}
	# POST genes to the /addList endpoint
	response = requests.post("%s/addList" % ENRICHR_URL, files=payload)
	list_ids = json.loads(response.text)
	return list_ids


def enrichr_link(genes, meta=''):
	"""POST a gene list to Enrichr server and get the link."""
	list_ids = _enrichr_add_list(genes, meta)
	shortId = list_ids['shortId']
	link = '%s/enrich?dataset=%s' % (ENRICHR_URL, shortId)
	return link


def enrichr_result(genes, meta='', gmt=''):
	"""POST the genes to Enrichr and return the enrichment results 
	for a specific gene-set library on Enrichr"""
	list_ids = _enrichr_add_list(genes, meta)
	# GET from the /export endpoint
	query_string = '?userListId=%s&backgroundType=%s' % \
		(list_ids['userListId'], gmt)

	url = '%s/enrich%s' % (ENRICHR_URL, query_string)
	sleep(2)

	response = requests.get(url)
	if response.status_code == 200:
		results = json.loads(response.text)
		return results
	else:
		raise Exception('HTTP reponse code=%s' % response.status_code)


def enrichr_term_score(genes, meta='', gmt=''):
	"""Use Enrichr API to only get terms and scores"""
	results = enrichr_result(genes, meta=meta, gmt=gmt)[gmt]
	terms_scores = []
	for res in results:
		term = res[1]
		combined_score = res[4]
		terms_scores.append((term, combined_score))
	return terms_scores

def cds2_link(chdir, meta='', aggravate=True):
	'''Use L1000CDS2 API to get a link for results'''
	url = 'http://amp.pharm.mssm.edu/L1000CDS2/query'
	data = {"genes": map(lambda x: x[1].upper(), chdir), "vals":map(lambda x: x[0], chdir)}
	config = {"aggravate":aggravate,"searchMethod":"CD","share":True,"combination":True,"db-version":"latest"}
	metadata = [{"key":"description","value": meta}]
	payload = {"data":data,"config":config,"meta":metadata}
	headers = {'content-type':'application/json'}
	r = requests.post(url,data=json.dumps(payload),headers=headers)
	resCD = r.json()
	shareId = resCD['shareId']
	result_url = 'http://amp.pharm.mssm.edu/L1000CDS2/#/result/' + shareId
	return result_url

def cds2_updn_link(upGenes, dnGenes, description='', aggravate=True):
	url = 'http://amp.pharm.mssm.edu/L1000CDS2/query'
	data = {"upGenes": map(lambda x: x.upper(), upGenes),
	"dnGenes":map(lambda x: x.upper(), dnGenes)}
	config = {"aggravate":aggravate,"searchMethod":"geneSet","share":True,"combination":True,"db-version":"latest"}
	metadata = [{"key":"Name","value":description}]
	payload = {"data":data,"config":config,"meta":metadata}
	headers = {'content-type':'application/json'}
	r = requests.post(url,data=json.dumps(payload),headers=headers)
	resGeneSet = r.json()
	print resGeneSet
	shareId = resGeneSet['shareId']
	result_url = 'http://amp.pharm.mssm.edu/L1000CDS2/#/result/' + shareId
	return result_url


def get_tracking_symbol(fn):
	with open (fn) as f:
		for line in f:
			sl = line.strip().split('\t')
			if ',' in sl[1]:
				d_tracking_gene[sl[0]] = sl[1].split(',')[0]
			else:
				d_tracking_gene[sl[0]] = sl[1]
	return d_tracking_gene

def parse_sigTable(fn, comparisons, d_tracking_gene=None):
	# parse the results from cummeRbund "getSigTable" method
	with open(fn) as f:
		header = next(f).strip().split('\t')
		index = [header.index(c)+1 for c in comparisons]
		d = {}
		for comp in comparisons:
			d[comp] = []
		for line in f:
			sl = line.strip().split('\t')
			if d_tracking_gene is not None:
				gene = d_tracking_gene[sl[0]]
			else:
				gene = sl[0]
			for comp, i in zip(comparisons, index):
				if sl[i] == '1':
					d[comp].append(gene)
	return d

def parse_fpkmMatrix(fn, d_tracking_gene):
	# parse fpkmMatrix with gene names
	d = {}
	with open (fn) as f:
		samples = next(f).strip().split('\t')
		for line in f:
			sl = line.strip().split('\t')
			gene = d_tracking_gene[sl[0]]
			vals = [float(s) for s in sl[1:]]
			d[gene] = vals
	return d, samples

def parse_DESeq(fn, padj_cutoff):
	# parse table written by DESeq nbinomTest function
	d = {} # gene : [l2fc, padj]
	with open (fn) as f:
		next(f)
		for line in f:
			sl = line.strip().split('\t')
			gene = sl[1]
			try:
				l2fc = float(sl[-3])
				padj = float(sl[-2])
				if padj < padj_cutoff:
					d[gene] = [l2fc, padj]
			except ValueError:
				pass
	return d

def fpkmMatrix2excel(fns, outfn):
	## take a list of fn(s) output by cummeRbund `fpkmMatrix` or `repFpkmMatrix`
	## with gene(s)|trackingID as row ids, and write into a Excel file
	if type(fns) != list:
		fns = [fns]
	book = xlwt.Workbook()
	for fn in fns:
		sheet = book.add_sheet(fn.split('.txt')[0])
		sheet.write(0,0,'gene')
		sheet.write(0,1,'gene_with_trackingID')
		with open (fn) as f:
			header = next(f).strip().split('\t')
			for i, item in enumerate(header):
				sheet.write(0, i+2, item)
			for row, line in enumerate(f,start=1):
				sl = line.strip().split('\t')
				gene_trackingID = sl[0]
				vals = sl[1:]
				if ',' in gene_trackingID:
					gene = gene_trackingID.split(',')[0]
				else:
					gene = gene_trackingID.split('|')[0]
				sheet.write(row, 0, gene)
				sheet.write(row, 1, gene_trackingID)
				for i, item in enumerate(vals,start=2):
					sheet.write(row, i, item)
	book.save(outfn)


def dict2xls(d, outfn, enrichr=True):
	book = xlwt.Workbook()
	sheet_links = book.add_sheet('Enrichr_links')
	sheet_links.write(0,0,'Gene list')
	sheet_links.write(0,1,'Size')
	sheet_links.write(0,2,'Link')
	keys = d.keys()
	keys.sort()
	for ik, key in enumerate(keys, start=1):
		sheet = book.add_sheet(key)
		genes = d[key]
		if enrichr:
			link = enrichr_link(genes, key)
		else:
			link = ''
		sheet_links.write(ik, 0, key)
		sheet_links.write(ik, 1, len(genes))
		sheet_links.write(ik, 2, link)
		for i, g in enumerate(genes):
			sheet.write(i, 0, g)
	book.save(outfn)
	return

def dict2xls_with_vals(d, colnames, outfn, enrichr=True):
	"""d is a dict of dict: {gene_list: {gene: [vals]} }
		colnames is a list of fieldnames correspond to the vals
	"""
	book = xlwt.Workbook()
	sheet_links = book.add_sheet('Enrichr_links')
	sheet_links.write(0,0,'Gene list')
	sheet_links.write(0,1,'Size')
	if enrichr:
		sheet_links.write(0,2,'Link')

	# check the value type of the dict inside d
	if not hasattr(d.values()[0].values()[0], '__len__'):
		# is scalar, convert to list
		for gene_list, dd in d.items():
			for gene, val in dd.items():
				dd[gene] = [val]
			d[gene_list] = dd
	
	gene_lists = d.keys()
	gene_lists.sort()
	for i, gene_list in enumerate(gene_lists, start=1):
		sheet = book.add_sheet(gene_list)
		genes = np.array(d[gene_list].keys())
		vals = np.array([vals[0] for vals in d[gene_list].values()])
		srt_idx = abs(vals).argsort()[::-1]
		genes = genes[srt_idx].tolist()
		if enrichr:
			link = enrichr_link(genes, gene_list)
			sheet_links.write(i, 2, link)
		sheet_links.write(i, 0, gene_list)
		sheet_links.write(i, 1, len(genes))
		

		for ic, col in enumerate(colnames): # write header for sheet
			sheet.write(0,ic,col)
		for ig, gene in enumerate(genes, start=1): # write gene and vals
			sheet.write(ig, 0, gene)
			vals = d[gene_list][gene]
			for iv, val in enumerate(vals, start=1):
				sheet.write(ig, iv, val)

	book.save(outfn)
	return


def dicts2xls2(d_up, d_dn, d_int, outfn):
	# write xls file of up/dn gene lists and their overlap
	# with lists of genes ls of interest
	# d_int should be a dict of lists
	book = xlwt.Workbook()
	sheet_links = book.add_sheet('Enrichr_links')
	sheet_links.write(0,0,'Gene list')
	sheet_links.write(0,1,'Size')
	sheet_links.write(0,2,'Link')
	for ik, key in enumerate(d_up, start=1):
		sheet = book.add_sheet(key + '_up')
		sheet.write(0, 0, 'All genes')
		genes = d_up[key]
		link = enrichr_link(genes, key)
		sheet_links.write(ik, 0, key)
		sheet_links.write(ik, 1, str(len(genes)))
		sheet_links.write(ik, 2, link)
		for i, g in enumerate(genes, start=1):
			sheet.write(i, 0, g)
		for i_n, l_name in enumerate(d_int, start=1): # gene list of interest
			genes_overlap = set(d_int[l_name]) & set(genes)
			sheet.write(0, i_n, 'Overlap with ' + l_name)
			for i, g in enumerate(genes_overlap, start=1):
				sheet.write(i, i_n, g)
	for ik, key in enumerate(d_dn, start=len(d_up)+1):
		sheet = book.add_sheet(key + '_dn')
		sheet.write(0, 0, 'All genes')
		genes = d_dn[key]
		link = enrichr_link(genes, key)
		sheet_links.write(ik, 0, key)
		sheet_links.write(ik, 1, str(len(genes)))
		sheet_links.write(ik, 2, link)
		for i, g in enumerate(genes, start=1):
			sheet.write(i, 0, g)
		for i_n, l_name in enumerate(d_int, start=1): # gene list of interest
			genes_overlap = set(d_int[l_name]) & set(genes)
			sheet.write(0, i_n, 'Overlap with ' + l_name)
			for i, g in enumerate(genes_overlap, start=1):
				sheet.write(i, i_n, g)

	book.save(outfn)
	return


def manifold_plot(man, fpkmMatrix, samples, standardize=3, log=True, show_text=False, sep='_', legend_loc='best', legend_size=14):
	# man: the instance of a manifold algorithm
	## preprocessing of the fpkmMatrix
	if log:
		fpkmMatrix = np.log10(fpkmMatrix + 1.)
	if standardize == 2: # standardize along rows/genes
		fpkmMatrix = zscore(fpkmMatrix, axis=1)
	elif standardize == 1: # standardize along cols/samples
		fpkmMatrix = zscore(fpkmMatrix, axis=0)

	fpkmMatrix = man.fit_transform(fpkmMatrix.T)
	fig = plt.figure(figsize=(8,8))
	ax = fig.add_subplot(111)
	scatter_proxies = []
	labels_show = []
	groups = {}
	conditions = list(set([s.split(sep)[0] for s in samples]))

	for row, label in zip(fpkmMatrix, samples):
		label_show = label.split(sep)[0]
		idx = conditions.index(label_show)
		ax.scatter(row[0], row[1], label='label', color=COLORS10[idx], visible=not show_text, s=50, marker='o')
		if label_show not in labels_show:
			labels_show.append(label_show)
			scatter1_proxy = Line2D([0],[0], ls="none", c=COLORS10[idx], marker='o')
			scatter_proxies.append(scatter1_proxy)
		if show_text:	
			ax.text(row[0], row[1], label, \
				ha='center', va='center', rotation=0, color=COLORS10[idx], size='large')
	
	ax.legend(scatter_proxies, labels_show, numpoints=1, frameon=True,loc=legend_loc, prop={'size':legend_size})
	ax.set_xlabel('M1', fontsize=20)
	ax.set_ylabel('M2', fontsize=20)
	enlarge_tick_fontsize(ax, 14)
	fig.tight_layout()
	plt.show()


	return


def perform_PCA(fpkmMatrix, standardize=3, log=True):
	## preprocessing of the fpkmMatrix
	if log:
		fpkmMatrix = np.log10(fpkmMatrix + 1.)
	if standardize == 2: # standardize along rows/genes
		fpkmMatrix = zscore(fpkmMatrix, axis=1)
	elif standardize == 1: # standardize along cols/samples
		fpkmMatrix = zscore(fpkmMatrix, axis=0)

	## remove genes with NaNs
	fpkmMatrix = fpkmMatrix[~np.isnan(np.sum(fpkmMatrix, axis=1))]

	pca = PCA(n_components=None)
	## get variance captured
	pca.fit(fpkmMatrix.T)
	variance_explained = pca.explained_variance_ratio_[0:3]
	variance_explained *= 100
	## compute PCA and plot
	pca_transformed = pca.transform(fpkmMatrix.T)
	return variance_explained, pca_transformed


def PCA_plot(fpkmMatrix, samples, standardize=3, log=True, show_text=False, sep='_', legend_loc='best', legend_size=14):
	# standardize: whether to a zscore transformation on the log10 transformed FPKM
	## perform PCA
	variance_explained, pca_transformed = perform_PCA(fpkmMatrix, standardize=standardize, log=log)

	fig = plt.figure(figsize=(8,8))
	ax = fig.add_subplot(111)
	scatter_proxies = []
	labels_show = []
	groups = {}
	conditions = list(set([s.split(sep)[0] for s in samples]))

	colors = COLORS10
	if len(conditions) > 10:
		colors = COLORS20
	if len(conditions) > 20:
		r = lambda: random.randint(0,255)
		colors = ['#%02X%02X%02X' % (r(),r(),r()) for i in range(len(conditions))]

	for row, label in zip(pca_transformed, samples):
		label_show = label.split(sep)[0]
		idx = conditions.index(label_show)
		ax.scatter(row[0], row[1], label='label', color=colors[idx], s=50, marker='o')
		if label_show not in labels_show:
			labels_show.append(label_show)
			scatter1_proxy = Line2D([0],[0], ls="none", c=colors[idx], marker='o')
			scatter_proxies.append(scatter1_proxy)
		if show_text:	
			ax.text(row[0], row[1]-2, label.split(sep)[1], \
				ha='center', va='center', rotation=0, color=colors[idx], size='large')
	
	ax.legend(scatter_proxies, labels_show, numpoints=1, frameon=True,loc=legend_loc, prop={'size':legend_size})
	ax.set_xlabel('PC1 (%.2f'%variance_explained[0] + '%' + ' variance captured)', fontsize=20)
	ax.set_ylabel('PC2 (%.2f'%variance_explained[1] + '%' + ' variance captured)', fontsize=20)
	enlarge_tick_fontsize(ax, 14)
	fig.tight_layout()
	plt.show()

def PCA_plot2(fpkmMatrix, color_by, shape_by, 
		standardize=3, log=True, legend_loc='best', legend_size=14):
	variance_explained, pca_transformed = perform_PCA(fpkmMatrix, standardize=standardize, log=log)

	fig = plt.figure(figsize=(8,8))
	ax = fig.add_subplot(111)
	scatter_proxies = []
	labels_show = [] # for legend and scatter proxies
	color_uniq = list(set(color_by))
	shape_uniq = list(set(shape_by))

	colors = COLORS10
	if len(color_uniq) > 10:
		colors = COLORS20
	if len(color_uniq) > 20:
		r = lambda: random.randint(0,255)
		colors = ['#%02X%02X%02X' % (r(),r(),r()) for i in range(len(color_uniq))]
	
	shapes = 'osv^phd'

	for row, label_c, label_s in zip(pca_transformed, color_by, shape_by):
		idx_c = color_uniq.index(label_c)
		idx_s = shape_uniq.index(label_s)
		ax.scatter(row[0], row[1], label='label', 
			color=colors[idx_c], s=50, marker=shapes[idx_s])
		label = '%s-%s' %(label_c, label_s)
		if label not in labels_show:
			labels_show.append(label)
			scatter1_proxy = Line2D([0],[0], ls="none", c=colors[idx_c], marker=shapes[idx_s])
			scatter_proxies.append(scatter1_proxy)

	ax.legend(scatter_proxies, labels_show, numpoints=1, frameon=True,loc=legend_loc, prop={'size':legend_size})
	ax.set_xlabel('PC1 (%.2f'%variance_explained[0] + '%' + ' variance captured)', fontsize=20)
	ax.set_ylabel('PC2 (%.2f'%variance_explained[1] + '%' + ' variance captured)', fontsize=20)
	enlarge_tick_fontsize(ax, 14)
	fig.tight_layout()
	plt.show()
	return

def PCA_3d_plot(fpkmMatrix, samples, standardize=3, log=True, show_text=False, sep='_', legend_loc='best', legend_size=14):
	# standardize: whether to a zscore transformation on the log10 transformed FPKM
	pca = PCA(n_components=None)
	## preprocessing of the fpkmMatrix
	if log:
		fpkmMatrix = np.log10(fpkmMatrix + 1.)	
	if standardize == 2: # standardize along rows/genes
		fpkmMatrix = zscore(fpkmMatrix, axis=1)
	elif standardize == 1: # standardize along cols/samples
		fpkmMatrix = zscore(fpkmMatrix, axis=0)
	
	## remove genes with NaNs
	fpkmMatrix = fpkmMatrix[~np.isnan(np.sum(fpkmMatrix, axis=1))]
	## get variance captured
	pca.fit(fpkmMatrix.T)
	variance_explained = pca.explained_variance_ratio_[0:3]
	variance_explained *= 100
	## compute PCA and plot
	pca = PCA(n_components=3)
	pca_transformed = pca.fit_transform(fpkmMatrix.T)
	fig = plt.figure(figsize=(9,9))
	ax = fig.add_subplot(111, projection='3d')
	labels_show = []
	scatter_proxies = []
	groups = {}
	conditions = list(set([s.split(sep)[0] for s in samples]))

	colors = COLORS10
	if len(conditions) > 10:
		colors = COLORS20
	if len(conditions) > 20:
		r = lambda: random.randint(0,255)
		colors = ['#%02X%02X%02X' % (r(),r(),r()) for i in range(len(conditions))]		

	for row, label in zip(pca_transformed, samples):
		label_show = label.split(sep)[0]
		idx = conditions.index(label_show)
		ax.scatter(row[0], row[1], row[2], label='label', color=colors[idx], s=50, marker='o')
		if label_show not in labels_show:
			labels_show.append(label_show)
			scatter1_proxy = Line2D([0],[0], ls="none", c=colors[idx], marker='o')
			scatter_proxies.append(scatter1_proxy)			
		if show_text:	
			ax.text(row[0], row[1]-5, row[2]-5, label.split(sep)[1], \
				ha='center', va='center', rotation=0, color=colors[idx], size='large')

	ax.set_xlabel('PC1 (%.2f'%variance_explained[0] + '%' + ' variance captured)', fontsize=16)
	ax.set_ylabel('PC2 (%.2f'%variance_explained[1] + '%' + ' variance captured)', fontsize=16)
	ax.set_zlabel('PC3 (%.2f'%variance_explained[2] + '%' + ' variance captured)', fontsize=16)
	ax.legend(scatter_proxies, labels_show, numpoints=1, frameon=True,loc='upper left',prop={'size':legend_size})
	fig.tight_layout()
	plt.show()


from statsmodels.sandbox.stats.multicomp import multipletests
from scipy.stats import f_oneway
def row_wise_anova(mat, categories, method='fdr_bh'):
	'''Apply one-way ANOVA to each row of mat, and adjust p-values.
	'''
	uniq_cats = np.unique(categories)
	pvals = np.ones(mat.shape[0], dtype=float)
	masks = [np.in1d(categories, [cat]) for cat in uniq_cats]
	for i in range(mat.shape[0]):
		row = mat[i]
		grouped_row = [ row[mask] for mask in masks ]
		fval, pval = f_oneway(*grouped_row)
		pvals[i] = pval

	_, qvals, _, _ = multipletests(pvals, method=method)
	return pvals, qvals

	